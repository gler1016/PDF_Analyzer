import pandas as pd
from pathlib import Path
from typing import List, Dict, Any
import logging
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

class ExcelExporter:
    def __init__(self):
        self.header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        self.header_font = Font(color="FFFFFF", bold=True)
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

    def export(self, contacts: List[Dict[str, Any]], output_path: Path):
        """
        Export contacts to an Excel file with proper formatting.
        
        Args:
            contacts: List of contact dictionaries to export
            output_path: Path where the Excel file should be saved
        """
        try:
            # Convert to DataFrame
            df = pd.DataFrame(contacts)
            
            # Reorder columns
            columns = [
                'company_name',
                'website',
                'contact_name',
                'job_title',
                'linkedin',
                'email',
                'source_pdf',
                'last_updated'
            ]
            df = df[columns]
            
            # Rename columns for display
            column_names = {
                'company_name': 'Company Name',
                'website': 'Company Website',
                'contact_name': 'Contact Person',
                'job_title': 'Job Title',
                'linkedin': 'LinkedIn Profile',
                'email': 'Email Address',
                'source_pdf': 'Source PDF',
                'last_updated': 'Last Updated'
            }
            df = df.rename(columns=column_names)
            
            # Save to Excel
            df.to_excel(output_path, index=False, engine='openpyxl')
            
            # Apply formatting
            self._format_excel(output_path)
            
            logging.info(f"Successfully exported {len(contacts)} contacts to {output_path}")
            
        except Exception as e:
            logging.error(f"Error exporting to Excel: {str(e)}")
            raise

    def _format_excel(self, file_path: Path):
        """Apply formatting to the Excel file."""
        from openpyxl import load_workbook
        
        wb = load_workbook(file_path)
        ws = wb.active
        
        # Format headers
        for cell in ws[1]:
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.border
        
        # Format data cells
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.border = self.border
                cell.alignment = Alignment(vertical='center')
        
        # Adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = min(adjusted_width, 50)
        
        # Freeze header row
        ws.freeze_panes = 'A2'
        
        # Save the formatted workbook
        wb.save(file_path) 